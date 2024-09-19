import streamlit as st
from openpyxl import load_workbook
from io import BytesIO

st.title("Excel 表格去除密码 App")

# File uploader widget
uploaded_file = st.file_uploader("请上传 Excel 文件", type="xlsx")

if uploaded_file is not None:
    # Get the original file name without the path
    original_filename = uploaded_file.name

    # Load the uploaded Excel file
    try:
        workbook = load_workbook(filename=BytesIO(uploaded_file.read()))
        for sheet in workbook.sheetnames:
            st.write(f"正在去除密码: {sheet}")
            workbook[sheet].protection.sheet = False

        # Save the unprotected file
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Modify the file name to include the prefix "无密码-"
        unprotected_filename = "无密码-" + original_filename
        
        # Provide the unprotected file for download
        st.download_button(label="下载去除密码的 Excel 文件", data=output, file_name=unprotected_filename)
        st.success("成功去除密码!")
    except Exception as e:
        st.error(f"Error processing file: {e}")
